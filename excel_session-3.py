# excel - real life applications
# import, process & export, pandas, matplotlib
import pandas as pd
import openpyxl as opx
import glob
#from openpyxl import load_workbook

#path = str(r'C: \Users\som87475\source\repos\python\coding club 2021\docs\Mott')
path = str('docs\Mott')
wb = opx.load_workbook(path + '\\' + 'May.xlsx')
wb.sheetnames
details = wb['Sheet1']
details = pd.DataFrame(details.values)
details.tail()
details.iloc[0:3]
print(details.iloc[0:3])
print(details.iloc[0:5,:]) # row by column
print(details.iloc[0:5, :2])
print(details.iloc[0:5, 2:4])

details.shape
details.info()
details['Quantity'].astype('int64').dtypes
details.describe()

details = details.iloc[1:]

writer = pd.ExcelWriter(path + '\\' + 'final_v2.xlsx')
details.to_excel(writer, index=False, sheet_name = 'new')
writer.save()
#details[details['Resource Name'] == 

#wb = opx.load_workbook(path + '\\' + 'final.xlsx')
wb['new'].column_dimensions['A'].width = 25
#wb['new'].column_dimensions['A'].lenght = 18

many_files = glob.glob('docs\Mott')
for each_file in many_files:
    print(each_file)
